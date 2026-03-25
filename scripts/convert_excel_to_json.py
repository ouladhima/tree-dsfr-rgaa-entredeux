#!/usr/bin/env python3
import argparse
import hashlib
import json
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

FALLBACK_PDF = "pdf/sample.pdf"

ALIASES = {
    "domaine_fonctionnel": {
        "domaine_fonctionnel",
        "domaine fonctionnel",
    },
    "famille": {
        "famille",
    },
    "intitule_er": {
        "intitule_emploi_reference_erxxxxxx",
        "intitule emploi reference erxxxxxx",
        "intitule_emploi_reference",
        "intitule emploi reference",
        "emploi_reference",
        "emploi reference",
        "er",
        "intitule_er",
    },
    "intitule_metier_fp": {
        "intitule_metier_fpxxxxxx",
        "intitule metier fpxxxxxx",
        "intitule_metier_fp",
        "intitule metier fp",
        "metier_fp",
        "metier fp",
        "fp",
    },
    "file_pdf": {
        "file_pdf",
        "file pdf",
        "pdf",
        "fichier_pdf",
        "fichier pdf",
    },
}

REQUIRED_KEYS = [
    "domaine_fonctionnel",
    "famille",
    "intitule_er",
    "intitule_metier_fp",
]


def normalize_header(value: str) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.lower()
    for token in ["(", ")", "[", "]"]:
        text = text.replace(token, " ")
    cleaned = []
    for char in text:
        if char.isalnum():
            cleaned.append(char)
        else:
            cleaned.append(" ")
    return "_".join("".join(cleaned).split())


def canonical_key(header: str) -> str | None:
    normalized = normalize_header(header)
    for key, aliases in ALIASES.items():
        normalized_aliases = {normalize_header(alias) for alias in aliases}
        if normalized in normalized_aliases:
            return key
    return None


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_leaf_id(parts: list[str]) -> str:
    digest = hashlib.sha1("||".join(parts).encode("utf-8")).hexdigest()[:12]
    return f"leaf-{digest}"


def get_or_create_child(parent: dict, name: str, level: str, path: list[str]) -> dict:
    parent.setdefault("_index", {})
    key = (level, name)

    if key not in parent["_index"]:
        digest = hashlib.sha1("||".join(path + [name]).encode("utf-8")).hexdigest()[:10]
        child = {
            "id": f"{level}-{digest}",
            "name": name,
            "level": level,
            "children": [],
        }
        parent["_index"][key] = child
        parent["children"].append(child)

    return parent["_index"][key]


def strip_indexes(node: dict) -> None:
    node.pop("_index", None)
    for child in node.get("children", []):
        strip_indexes(child)


def compute_leaf_count(node: dict) -> int:
    children = node.get("children", [])
    if not children:
        node["leafCount"] = 1
        return 1

    total = sum(compute_leaf_count(child) for child in children)
    node["leafCount"] = total
    return total


def read_rows(workbook_path: Path) -> tuple[list[dict], int]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Le fichier Excel est vide.")

    raw_headers = [normalize_cell(cell) for cell in rows[0]]
    canonical_headers = [canonical_key(header) for header in raw_headers]

    missing = [key for key in REQUIRED_KEYS if key not in canonical_headers]
    if missing:
        raise ValueError(
            "Colonnes manquantes dans le fichier Excel : "
            + ", ".join(missing)
            + ". Vérifiez les intitulés de colonnes."
        )

    extracted_rows = []
    skipped = 0

    for excel_row in rows[1:]:
        values = {canonical_headers[index]: normalize_cell(value) for index, value in enumerate(excel_row) if index < len(canonical_headers) and canonical_headers[index]}
        required_values = [values.get(key, "") for key in REQUIRED_KEYS]

        if not any(required_values):
            continue

        if not all(required_values):
            skipped += 1
            continue

        extracted_rows.append(
            {
                "domaine_fonctionnel": values["domaine_fonctionnel"],
                "famille": values["famille"],
                "intitule_er": values["intitule_er"],
                "intitule_metier_fp": values["intitule_metier_fp"],
                "file_pdf": values.get("file_pdf", "").strip(),
            }
        )

    return extracted_rows, skipped


def build_output(rows: list[dict], source_file: str, skipped_rows: int, fallback_pdf: str) -> dict:
    root = {
        "id": "root",
        "name": "Cartographie RMFP",
        "level": "root",
        "children": [],
    }
    records = []

    for row in rows:
        domain = get_or_create_child(root, row["domaine_fonctionnel"], "domaine", [root["name"]])
        family = get_or_create_child(domain, row["famille"], "famille", [root["name"], row["domaine_fonctionnel"]])
        emploi = get_or_create_child(
            family,
            row["intitule_er"],
            "emploi_reference",
            [root["name"], row["domaine_fonctionnel"], row["famille"]],
        )

        leaf_path = [
            root["name"],
            row["domaine_fonctionnel"],
            row["famille"],
            row["intitule_er"],
            row["intitule_metier_fp"],
        ]
        leaf_id = build_leaf_id(leaf_path)

        leaf = {
            "id": leaf_id,
            "name": row["intitule_metier_fp"],
            "level": "metier_fp",
            "file_pdf": row["file_pdf"] or fallback_pdf,
            "path": leaf_path,
        }
        emploi["children"].append(leaf)

        records.append(
            {
                "id": leaf_id,
                "domaine_fonctionnel": row["domaine_fonctionnel"],
                "famille": row["famille"],
                "intitule_er": row["intitule_er"],
                "intitule_metier_fp": row["intitule_metier_fp"],
                "file_pdf": row["file_pdf"] or fallback_pdf,
                "path": leaf_path,
            }
        )

    strip_indexes(root)
    compute_leaf_count(root)

    return {
        "fallbackPdf": fallback_pdf,
        "meta": {
            "source_file": source_file,
            "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
            "total_records": len(records),
            "skipped_rows": skipped_rows,
            "levels": [
                "Domaine fonctionnel",
                "Famille",
                "Intitulé Emploi Référence",
                "Intitulé Métier FP",
            ],
        },
        "tree": root,
        "records": records,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convertit le fichier Excel RMFP en JSON exploitable par le site statique."
    )
    parser.add_argument(
        "input_xlsx",
        nargs="?",
        default="Correspondance RMFP.xlsx",
        help="Chemin vers le fichier Excel source (défaut: Correspondance RMFP.xlsx).",
    )
    parser.add_argument(
        "output_json",
        nargs="?",
        default="data/rmfp-data.json",
        help="Chemin de sortie du JSON généré (défaut: data/rmfp-data.json).",
    )
    parser.add_argument(
        "--fallback-pdf",
        default=FALLBACK_PDF,
        help=f"PDF de secours utilisé si la colonne file_pdf est vide (défaut: {FALLBACK_PDF}).",
    )
    args = parser.parse_args()

    input_path = Path(args.input_xlsx)
    output_path = Path(args.output_json)

    if not input_path.exists():
        print(f"[ERREUR] Fichier introuvable : {input_path}", file=sys.stderr)
        return 1

    try:
        rows, skipped_rows = read_rows(input_path)
        payload = build_output(rows, input_path.name, skipped_rows, args.fallback_pdf)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        print(f"[ERREUR] {exc}", file=sys.stderr)
        return 1

    print(
        f"[OK] JSON généré : {output_path} | "
        f"{payload['meta']['total_records']} enregistrement(s), "
        f"{payload['meta']['skipped_rows']} ligne(s) ignorée(s)."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
